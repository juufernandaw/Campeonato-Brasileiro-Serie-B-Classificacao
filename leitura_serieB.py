import operator
import openpyxl
import pandas as pd

# load excel with its path
wrkbk = openpyxl.load_workbook("tabela_serieB.xlsx")
todas_rodadas = []
sh = wrkbk.active
cont = -1
for sh in wrkbk.worksheets:
    uma_rodada = []
    cont += 1
    for i in range(2, sh.max_row + 1):

        rodada = sh.cell(row=i, column=2).value
        mandante = sh.cell(row=i, column=3).value
        visitante = sh.cell(row=i, column=4).value
        placar_mandante = int(sh.cell(row=i, column=5).value)
        placar_visitante = int(sh.cell(row=i, column=6).value)
        if placar_mandante == placar_visitante:     # se for empate
            total_pontos_mandante = total_pontos_visitante = 1
        elif placar_mandante > placar_visitante:    # se o mandante vencer
            total_pontos_mandante = 3
            total_pontos_visitante = 0
        else:   # se o mandante perder
            total_pontos_mandante = 0
            total_pontos_visitante = 3
        if cont > 0:
            # iterate just once and then save the found value; why? to optimize
            time_mandante = next(item for item in todas_rodadas[cont-1] if item["time"] == mandante)
            time_visitante = next(item for item in todas_rodadas[cont-1] if item["time"] == visitante)
            saldo_mandante = time_mandante["saldo_gols"]
            saldo_visitante = time_visitante["saldo_gols"]
            gols_pro_mandante = time_mandante["gols_pro"]
            gols_pro_visitante = time_visitante["gols_pro"]
            gols_contra_mandante = time_mandante["gols_contra"]
            gols_contra_visitante = time_visitante["gols_contra"]
            pontos_mandante = time_mandante["total_pontos"]
            pontos_visitante = time_visitante["total_pontos"]
            vitorias_mandante = time_mandante["vitorias"]
            vitorias_visitante = time_visitante["vitorias"]
        else:
            saldo_mandante = saldo_visitante = 0
            pontos_mandante = pontos_visitante = gols_pro_mandante = gols_pro_visitante = gols_contra_mandante = gols_contra_visitante = vitorias_mandante = vitorias_visitante = 0
        vitorias_mandante = vitorias_mandante + 1 if total_pontos_mandante == 3 else vitorias_mandante
        vitorias_visitante = vitorias_visitante + 1 if total_pontos_visitante == 3 else vitorias_visitante
        classificacao_mandante = {'rodada': rodada, 'time': mandante, 'total_pontos': pontos_mandante+total_pontos_mandante, 'gols_pro': gols_pro_mandante+placar_mandante, 'gols_contra': gols_contra_mandante+placar_visitante, 'saldo_gols': (saldo_mandante+placar_mandante)-placar_visitante, 'vitorias': vitorias_mandante}
        classificacao_visitante = {'rodada': rodada, 'time': visitante, 'total_pontos': pontos_visitante+total_pontos_visitante, 'gols_pro': gols_pro_visitante+placar_visitante, 'gols_contra': gols_contra_visitante+placar_mandante, 'saldo_gols': (saldo_visitante+placar_visitante)-placar_mandante, 'vitorias': vitorias_visitante}
        uma_rodada.append(classificacao_visitante)
        uma_rodada.append(classificacao_mandante)
    todas_rodadas.append(uma_rodada)

writer = pd.ExcelWriter('classificacao_por_rodada.xlsx')
for num_rodada, rodada in enumerate(todas_rodadas):
    rodada_ordenada = sorted(rodada, key=operator.itemgetter('total_pontos', 'vitorias', 'saldo_gols', 'gols_pro'), reverse=True)
    df_rodadas = pd.DataFrame(data=rodada_ordenada)

    df_rodadas.to_excel(writer, sheet_name=f'Rodada {num_rodada+1}')
writer.save()

